from typing import List, Optional
from decimal import Decimal
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import extract, func
from app.database import get_db
from app import models, schemas

router = APIRouter(prefix="/api/giving", tags=["Giving"])


@router.get("", response_model=List[schemas.GivingOut])
def list_giving(
    year:      Optional[int] = Query(None),
    member_id: Optional[str] = Query(None),
    type:      Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(models.GivingRecord).options(joinedload(models.GivingRecord.member))
    if year:
        q = q.filter(extract('year', models.GivingRecord.date) == year)
    if member_id:
        q = q.filter(models.GivingRecord.member_id == member_id)
    if type:
        q = q.filter(models.GivingRecord.type == type)
    records = q.order_by(models.GivingRecord.date.desc()).all()
    result = []
    for r in records:
        out = schemas.GivingOut.from_orm(r)
        out.member_name = f"{r.member.first} {r.member.last}" if r.member else "Anonymous"
        result.append(out)
    return result


@router.get("/summary")
def giving_summary(year: Optional[int] = Query(None), db: Session = Depends(get_db)):
    """Returns total giving by fund/type for the given year (defaults to current year)."""
    from datetime import datetime
    y = year or datetime.utcnow().year
    rows = (
        db.query(models.GivingRecord.type, func.sum(models.GivingRecord.amount).label("total"))
        .filter(extract('year', models.GivingRecord.date) == y)
        .group_by(models.GivingRecord.type)
        .all()
    )
    return {r.type: float(r.total) for r in rows}


@router.post("", response_model=schemas.GivingOut, status_code=201)
def create_giving(data: schemas.GivingCreate, db: Session = Depends(get_db)):
    record = models.GivingRecord(**data.dict())
    db.add(record)
    db.commit()
    db.refresh(record)
    out = schemas.GivingOut.from_orm(record)
    if record.member:
        out.member_name = f"{record.member.first} {record.member.last}"
    return out


@router.delete("/{giving_id}", status_code=204)
def delete_giving(giving_id: str, db: Session = Depends(get_db)):
    r = db.query(models.GivingRecord).filter(models.GivingRecord.id == giving_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Record not found")
    db.delete(r)
    db.commit()


# ── Export / Import ───────────────────────────────────────────────────────────

GIVING_FIELDS = [
    'date', 'amount', 'type', 'fund', 'notes',
    'member_email', 'member_first', 'member_last',
]

def _giving_row(g):
    return [
        str(g.date) if g.date else '',
        str(g.amount),
        g.type or '',
        g.fund or 'General Fund',
        g.notes or '',
        (g.member.email or '') if g.member else '',
        (g.member.first or '') if g.member else '',
        (g.member.last or '') if g.member else '',
    ]


@router.get("/export")
def export_giving(
    format: str = Query("csv"),
    year: Optional[int] = Query(None),
    member_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    import io, csv as _csv
    from fastapi.responses import StreamingResponse, Response

    q = db.query(models.GivingRecord).options(joinedload(models.GivingRecord.member))
    if year:
        q = q.filter(extract('year', models.GivingRecord.date) == year)
    if member_id:
        q = q.filter(models.GivingRecord.member_id == member_id)
    records = q.order_by(models.GivingRecord.date.desc()).all()

    if format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed — run: pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Giving"
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(GIVING_FIELDS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for g in records:
            ws.append(_giving_row(g))
        # Format amount column (B) as currency
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '"$"#,##0.00'
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"giving_{year or 'all'}.xlsx"
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    delimiter = '\t' if format == 'tsv' else ','
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter)
    writer.writerow(GIVING_FIELDS)
    for g in records:
        writer.writerow(_giving_row(g))
    ext = 'tsv' if format == 'tsv' else 'csv'
    mime = 'text/tab-separated-values' if format == 'tsv' else 'text/csv'
    fname = f"giving_{year or 'all'}.{ext}"
    return Response(content=buf.getvalue(), media_type=mime,
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/template")
def giving_template(format: str = Query("csv")):
    import io, csv as _csv
    from fastapi.responses import StreamingResponse, Response

    example = ['2024-01-07', '100.00', 'Tithe', 'General Fund', '', 'jane@example.com', 'Jane', 'Smith']

    if format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Giving"
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(GIVING_FIELDS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.append(example)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="giving_template.xlsx"'})

    delimiter = '\t' if format == 'tsv' else ','
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter)
    writer.writerow(GIVING_FIELDS)
    writer.writerow(example)
    ext = 'tsv' if format == 'tsv' else 'csv'
    mime = 'text/tab-separated-values' if format == 'tsv' else 'text/csv'
    return Response(content=buf.getvalue(), media_type=mime,
                    headers={"Content-Disposition": f'attachment; filename="giving_template.{ext}"'})


@router.post("/import")
async def import_giving(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    import io, csv as _csv
    from decimal import Decimal as _Dec
    from datetime import datetime as _dt

    content = await file.read()
    fname = (file.filename or '').lower()

    rows = []
    if fname.endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h or '').strip().lower() for h in raw_headers]
        for row_idx in range(2, ws.max_row + 1):
            vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == '' for v in vals):
                continue
            rows.append({headers[i]: (str(v).strip() if v is not None else '')
                         for i, v in enumerate(vals)})
    else:
        delimiter = '\t' if (fname.endswith('.tsv') or fname.endswith('.txt')) else ','
        text = content.decode('utf-8-sig', errors='replace')
        reader = _csv.DictReader(io.StringIO(text), delimiter=delimiter)
        norm_headers = [f.strip().lower() for f in (reader.fieldnames or [])]
        reader.fieldnames = norm_headers
        for r in reader:
            rows.append({k.strip().lower(): (v or '').strip() for k, v in r.items()})

    created, skipped, errors = 0, 0, []

    for i, r in enumerate(rows, 2):
        date_str   = r.get('date', '').strip()
        amount_str = r.get('amount', '').strip().lstrip('$').replace(',', '')

        if not date_str or not amount_str:
            errors.append(f"Row {i}: missing date or amount — skipped")
            continue

        # Parse date
        rec_date = None
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
            try:
                rec_date = _dt.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue
        if not rec_date:
            errors.append(f"Row {i}: unrecognised date '{date_str}' — skipped")
            continue

        # Parse amount
        try:
            amount = _Dec(amount_str)
        except Exception:
            errors.append(f"Row {i}: invalid amount '{amount_str}' — skipped")
            continue

        fund   = r.get('fund', '').strip() or 'General Fund'
        gtype  = r.get('type', '').strip() or 'Tithe'
        notes  = r.get('notes', '').strip() or None

        # Resolve member by email
        member_id = None
        email_str = r.get('member_email', '').strip().lower()
        if email_str:
            mem = db.query(models.Member).filter(
                models.Member.email.ilike(email_str)).first()
            if mem:
                member_id = mem.id

        # Duplicate check: same date + amount + fund + member
        existing = db.query(models.GivingRecord).filter(
            models.GivingRecord.date == rec_date,
            models.GivingRecord.amount == amount,
            models.GivingRecord.fund == fund,
            models.GivingRecord.member_id == member_id,
        ).first()
        if existing:
            skipped += 1
            continue

        db.add(models.GivingRecord(
            date=rec_date, amount=amount,
            type=gtype, fund=fund, notes=notes,
            member_id=member_id,
        ))
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
