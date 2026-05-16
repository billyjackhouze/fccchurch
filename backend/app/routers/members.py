from typing import List, Optional
import os, shutil
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app import models, schemas

router = APIRouter(prefix="/api/members", tags=["Members"])

RELATION_REVERSE = {
    'Partner':    'Partner',
    'Child':      'Parent',
    'Parent':     'Child',
    'Sibling':    'Sibling',
    'Guardian':   'Ward',
    'Ward':       'Guardian',
    'Grandparent':'Grandchild',
    'Grandchild': 'Grandparent',
    'Aunt/Uncle': 'Niece/Nephew',
    'Niece/Nephew':'Aunt/Uncle',
    'Other':      'Other',
}


def enrich_member(m: models.Member) -> schemas.MemberOut:
    out = schemas.MemberOut.from_orm(m)
    family = []
    for r in m.relationships_from:
        if r.related:
            family.append(schemas.MemberRelationshipOut(
                id=r.id,
                related_id=r.related_id,
                related_name=f"{r.related.first} {r.related.last}",
                relation=r.relation,
                related_photo=r.related.photo,
            ))
    out.family = family
    return out


def q_members(db):
    return db.query(models.Member).options(
        joinedload(models.Member.relationships_from).joinedload(models.MemberRelationship.related)
    )


@router.get("", response_model=List[schemas.MemberOut])
def list_members(
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = q_members(db)
    if status:
        q = q.filter(models.Member.status == status)
    if search:
        term = f"%{search}%"
        q = q.filter(
            models.Member.first.ilike(term) |
            models.Member.last.ilike(term)  |
            models.Member.email.ilike(term) |
            models.Member.ministry.ilike(term)
        )
    return [enrich_member(m) for m in q.order_by(models.Member.last, models.Member.first).all()]


@router.post("", response_model=schemas.MemberOut, status_code=201)
def create_member(data: schemas.MemberCreate, db: Session = Depends(get_db)):
    member = models.Member(**data.dict())
    db.add(member)
    db.commit()
    db.refresh(member)
    return enrich_member(member)


@router.get("/{member_id}", response_model=schemas.MemberOut)
def get_member(member_id: str, db: Session = Depends(get_db)):
    m = q_members(db).filter(models.Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    return enrich_member(m)


@router.put("/{member_id}", response_model=schemas.MemberOut)
def update_member(member_id: str, data: schemas.MemberUpdate, db: Session = Depends(get_db)):
    m = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    for k, v in data.dict(exclude_unset=True).items():
        setattr(m, k, v)
    db.commit()
    return enrich_member(q_members(db).filter(models.Member.id == member_id).first())


@router.delete("/{member_id}", status_code=204)
def delete_member(member_id: str, db: Session = Depends(get_db)):
    m = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    db.delete(m)
    db.commit()


# ── Photo Upload ──────────────────────────────────────────────────────────────

@router.post("/{member_id}/photo")
async def upload_photo(member_id: str, file: UploadFile = File(...),
                       db: Session = Depends(get_db)):
    m = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")

    ext = (file.filename or "photo.jpg").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
        ext = "jpg"
    filename = f"{member_id}.{ext}"

    photos_dir = os.path.join(os.path.dirname(__file__), "..", "..", "static", "photos")
    os.makedirs(photos_dir, exist_ok=True)

    # Remove old photo if different extension
    for old_ext in ("jpg", "jpeg", "png", "gif", "webp"):
        old_path = os.path.join(photos_dir, f"{member_id}.{old_ext}")
        if os.path.exists(old_path) and old_ext != ext:
            os.remove(old_path)

    file_path = os.path.join(photos_dir, filename)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    m.photo = filename
    db.commit()
    return {"photo": filename, "url": f"/static/photos/{filename}"}


# ── Family Relationships ──────────────────────────────────────────────────────

@router.post("/{member_id}/family", response_model=schemas.MemberRelationshipOut, status_code=201)
def add_family_member(member_id: str, data: schemas.MemberRelationshipCreate,
                      db: Session = Depends(get_db)):
    if not db.query(models.Member).filter(models.Member.id == member_id).first():
        raise HTTPException(status_code=404, detail="Member not found")
    related = db.query(models.Member).filter(models.Member.id == data.related_id).first()
    if not related:
        raise HTTPException(status_code=404, detail="Related member not found")

    rel = models.MemberRelationship(member_id=member_id, related_id=data.related_id, relation=data.relation)
    rev = models.MemberRelationship(member_id=data.related_id, related_id=member_id,
                                    relation=RELATION_REVERSE.get(data.relation, "Other"))
    db.add(rel); db.add(rev)
    db.commit(); db.refresh(rel)

    return schemas.MemberRelationshipOut(
        id=rel.id, related_id=rel.related_id,
        related_name=f"{related.first} {related.last}",
        relation=rel.relation,
        related_photo=related.photo,
    )


# ── Member Activity (giving + pledges summary) ───────────────────────────────

@router.get("/{member_id}/activity")
def get_member_activity(member_id: str, db: Session = Depends(get_db)):
    from datetime import date as dt_date
    m = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Member not found")
    now_year = dt_date.today().year
    giving_records = db.query(models.GivingRecord).filter(
        models.GivingRecord.member_id == member_id
    ).order_by(models.GivingRecord.date.desc()).all()
    pledges = db.query(models.Pledge).filter(
        models.Pledge.member_id == member_id
    ).all()
    giving_total      = float(sum(g.amount for g in giving_records))
    giving_this_year  = float(sum(g.amount for g in giving_records if g.date.year == now_year))
    return {
        "giving": [{"id":g.id,"date":str(g.date),"amount":float(g.amount),"type":g.type,"fund":g.fund,"notes":g.notes} for g in giving_records],
        "pledges": [{"id":p.id,"campaign":p.campaign,"pledged_amount":float(p.pledged_amount),"paid_amount":float(p.paid_amount),"pledge_date":str(p.pledge_date) if p.pledge_date else None,"end_date":str(p.end_date) if p.end_date else None,"frequency":p.frequency,"status":p.status} for p in pledges],
        "giving_total": giving_total,
        "giving_this_year": giving_this_year,
    }


@router.delete("/{member_id}/family/{relation_id}", status_code=204)
def remove_family_member(member_id: str, relation_id: str, db: Session = Depends(get_db)):
    rel = db.query(models.MemberRelationship).filter(
        models.MemberRelationship.id == relation_id).first()
    if not rel:
        raise HTTPException(status_code=404, detail="Relationship not found")
    reverse = db.query(models.MemberRelationship).filter(
        models.MemberRelationship.member_id == rel.related_id,
        models.MemberRelationship.related_id == rel.member_id
    ).first()
    db.delete(rel)
    if reverse:
        db.delete(reverse)
    db.commit()


# ── Export / Import ───────────────────────────────────────────────────────────

MEMBER_FIELDS = [
    'first', 'last', 'email', 'phone', 'address',
    'status', 'since', 'ministry', 'family_size', 'pronouns', 'notes',
]

def _member_row(m):
    return [
        m.first or '', m.last or '', m.email or '', m.phone or '',
        m.address or '', m.status or 'Active',
        str(m.since) if m.since else '',
        m.ministry or '', m.family_size or 1,
        m.pronouns or '', m.notes or '',
    ]


@router.get("/export")
def export_members(
    format: str = Query("csv"),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    import io, csv as _csv
    from fastapi.responses import StreamingResponse, Response

    q = q_members(db)
    if status:
        q = q.filter(models.Member.status == status)
    if search:
        term = f"%{search}%"
        q = q.filter(
            models.Member.first.ilike(term) | models.Member.last.ilike(term) |
            models.Member.email.ilike(term) | models.Member.ministry.ilike(term)
        )
    members = q.order_by(models.Member.last, models.Member.first).all()

    if format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed — run: pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(MEMBER_FIELDS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for m in members:
            ws.append(_member_row(m))
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="members.xlsx"'})

    delimiter = '\t' if format == 'tsv' else ','
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter)
    writer.writerow(MEMBER_FIELDS)
    for m in members:
        writer.writerow(_member_row(m))
    ext = 'tsv' if format == 'tsv' else 'csv'
    mime = 'text/tab-separated-values' if format == 'tsv' else 'text/csv'
    return Response(content=buf.getvalue(), media_type=mime,
                    headers={"Content-Disposition": f'attachment; filename="members.{ext}"'})


@router.get("/template")
def member_template(format: str = Query("csv")):
    import io, csv as _csv
    from fastapi.responses import StreamingResponse, Response

    if format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(MEMBER_FIELDS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # Example row
        ws.append(['Jane', 'Smith', 'jane@example.com', '555-1234',
                   '123 Main St', 'Active', '2020-01-15', 'Worship', 2, 'she/her', ''])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="members_template.xlsx"'})

    delimiter = '\t' if format == 'tsv' else ','
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter)
    writer.writerow(MEMBER_FIELDS)
    writer.writerow(['Jane', 'Smith', 'jane@example.com', '555-1234',
                     '123 Main St', 'Active', '2020-01-15', 'Worship', 2, 'she/her', ''])
    ext = 'tsv' if format == 'tsv' else 'csv'
    mime = 'text/tab-separated-values' if format == 'tsv' else 'text/csv'
    return Response(content=buf.getvalue(), media_type=mime,
                    headers={"Content-Disposition": f'attachment; filename="members_template.{ext}"'})


@router.post("/import")
async def import_members(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    import io, csv as _csv
    from datetime import datetime as _dt

    content = await file.read()
    fname = (file.filename or '').lower()

    rows = []
    if fname.endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h or '').strip().lower() for h in raw_headers]
        for row_idx in range(2, ws.max_row + 1):
            vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == '' for v in vals):
                continue
            rows.append({headers[i]: (str(v).strip() if v is not None else '')
                         for i, v in enumerate(vals)})
    else:
        delimiter = '\t' if (fname.endswith('.tsv') or fname.endswith('.txt')) else ','
        text = content.decode('utf-8-sig', errors='replace')
        reader = _csv.DictReader(io.StringIO(text), delimiter=delimiter)
        norm_headers = [f.strip().lower() for f in (reader.fieldnames or [])]
        reader.fieldnames = norm_headers
        for r in reader:
            rows.append({k.strip().lower(): (v or '').strip() for k, v in r.items()})

    created, skipped, errors = 0, 0, []

    for i, r in enumerate(rows, 2):
        first = r.get('first', '').strip()
        last  = r.get('last', '').strip()
        email = (r.get('email', '').strip().lower()) or None

        if not first or not last:
            errors.append(f"Row {i}: missing first or last name — skipped")
            continue

        # Duplicate check by email
        if email:
            existing = db.query(models.Member).filter(
                models.Member.email.ilike(email)).first()
            if existing:
                skipped += 1
                continue

        # Parse member-since date
        since = None
        since_str = r.get('since', '').strip()
        if since_str:
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d', '%d-%m-%Y'):
                try:
                    since = _dt.strptime(since_str, fmt).date()
                    break
                except ValueError:
                    continue

        family_size = 1
        try:
            family_size = int(float(r.get('family_size', '1') or '1'))
        except (ValueError, TypeError):
            pass

        db.add(models.Member(
            first=first, last=last, email=email,
            phone=r.get('phone') or None,
            address=r.get('address') or None,
            status=r.get('status') or 'Active',
            since=since,
            ministry=r.get('ministry') or None,
            family_size=family_size,
            pronouns=r.get('pronouns') or None,
            notes=r.get('notes') or None,
        ))
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
