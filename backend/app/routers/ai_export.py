"""
AI-powered natural language export tool (admin only).

Flow:
  1. POST /preview  — user sends plain-English request
                      → Claude reads schema, generates SQL
                      → we validate (SELECT-only), execute, return first 20 rows + full SQL
  2. POST /download — client sends back the validated SQL + desired format
                      → we re-validate and stream the file (CSV / TSV / XLSX)
"""
import re
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.database import get_db, engine
from app import models
from app.routers.auth import require_admin
from app.routers.settings import get_raw

router = APIRouter(prefix="/api/ai-export", tags=["AI Export"])

# ── Schema description sent to Claude ─────────────────────────────────────────
SCHEMA = """
PostgreSQL schema — FFC Church Management System
(All PKs are text UUID strings)

members            : id, first, last, email, phone, address,
                     status[Active|Inactive|Guest|Visitor|Former Member],
                     since(date), ministry(text), family_size(int),
                     pronouns, notes, created_at

giving             : id, member_id→members.id(nullable),
                     date, amount(numeric), type(text), fund(text), notes, created_at
                     Common types: Tithe, Offering, Special Gift, Building Fund, Missions

pledges            : id, member_id→members.id(nullable), campaign,
                     pledged_amount(numeric), paid_amount(numeric),
                     pledge_date, end_date, frequency, status, created_at

events             : id, title, date, start_time, end_time, type,
                     room_id→rooms.id(nullable), organizer, description,
                     volunteer_slots(int), created_at

rooms              : id, name, capacity(int), location, features, notes, color, created_at

ministries         : id, name, description, leader_id→members.id(nullable),
                     color, created_at

ministry_memberships: id, ministry_id→ministries.id, member_id→members.id,
                     role, joined_date, created_at

church_groups      : id, name, group_type, leader_id→members.id(nullable),
                     meeting_day, meeting_time, location, description,
                     is_active(boolean), color, created_at

group_memberships  : id, group_id→church_groups.id, member_id→members.id,
                     role, joined_date, created_at

volunteer_shifts   : id, title, ministry, date, start_time, end_time,
                     room_id→rooms.id(nullable), description,
                     slots_needed(int), created_at

shift_signups      : id, shift_id→volunteer_shifts.id,
                     member_id→members.id, signed_up_at

service_plans      : id, title, date, service_type,
                     status[draft|planning|ready|complete],
                     series_name, sermon_title, sermon_scripture,
                     preacher_id→members.id(nullable), notes, created_at

sermons            : id, title, date, series_name, scripture,
                     preacher_id→members.id(nullable),
                     sermon_notes, tags, created_at
"""


# ── Safety ────────────────────────────────────────────────────────────────────
_WRITE_OPS = re.compile(
    r'\b(INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|TRUNCATE|EXEC(?:UTE)?|GRANT|REVOKE|REPLACE|MERGE)\b',
    re.IGNORECASE,
)

def _validate_select(sql: str) -> bool:
    """Return True only if sql is a safe, read-only SELECT statement."""
    cleaned = sql.strip().rstrip(';').lstrip()
    upper   = cleaned.upper()
    if not (upper.startswith('SELECT') or upper.startswith('WITH')):
        return False
    if _WRITE_OPS.search(cleaned):
        return False
    return True


def _strip_fences(text: str) -> str:
    """Remove markdown code fences if Claude wrapped the SQL in them."""
    text = text.strip()
    if text.startswith('```'):
        lines = text.split('\n')
        text  = '\n'.join(lines[1:])
    if text.endswith('```'):
        text = '\n'.join(text.split('\n')[:-1])
    return text.strip().rstrip(';')


# ── Request / Response models ─────────────────────────────────────────────────
class PreviewRequest(BaseModel):
    query: str                       # Natural-language export request
    row_limit: Optional[int] = 1000  # Max rows in full download


class DownloadRequest(BaseModel):
    sql:    str                      # Validated SQL from a previous /preview call
    format: Optional[str] = 'csv'   # csv | tsv | xlsx


# ── /preview ──────────────────────────────────────────────────────────────────
@router.post('/preview')
def preview_export(req: PreviewRequest,
                   db: Session = Depends(get_db),
                   _:  models.User = Depends(require_admin)):
    """
    1. Send NL query + schema to Claude → get SQL
    2. Validate SELECT-only
    3. Execute → return first 20 rows as JSON, plus full SQL and row count
    """
    api_key = get_raw('api_anthropic_key', db)
    if not api_key:
        raise HTTPException(status_code=400,
            detail='Anthropic API key not configured. Add it in Admin → Settings → API Keys.')

    try:
        import anthropic
    except ImportError:
        raise HTTPException(status_code=500, detail='anthropic library not installed.')

    limit  = min(req.row_limit or 1000, 5000)
    today  = date.today().isoformat()

    prompt = f"""You are a PostgreSQL query generator for a church management database.

SCHEMA:
{SCHEMA}

TODAY: {today}

REQUEST: "{req.query}"

Generate a single valid PostgreSQL SELECT query that fulfills this request.

Rules — strictly follow every one:
1. Return ONLY the raw SQL query — no explanation, no markdown, no code fences, no comments
2. Use only SELECT statements — never INSERT, UPDATE, DELETE, DROP, or any write operation
3. Give every column a clear alias (e.g.  m.first || ' ' || m.last  AS member_name)
4. Include LIMIT {limit} unless the request explicitly asks for all records
5. JOIN related tables as needed; use LEFT JOIN when data may be absent
6. Wrap amount/money columns with  ROUND(col::numeric, 2)
7. For "this year" use  EXTRACT(year FROM col) = EXTRACT(year FROM CURRENT_DATE)
8. For "last N months" use  col >= CURRENT_DATE - INTERVAL 'N months'
9. Order results sensibly (alphabetical by name, or date DESC for time-based data)
10. Use COALESCE for nullable text columns where appropriate

Return ONLY the SQL — nothing else."""

    client  = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-opus-4-6',
        max_tokens=1000,
        messages=[{'role': 'user', 'content': prompt}],
    )

    raw_sql = _strip_fences(message.content[0].text)

    if not _validate_select(raw_sql):
        raise HTTPException(status_code=400,
            detail=f'AI returned an unsafe query. Please rephrase your request.\n\nRaw output:\n{raw_sql}')

    # Execute
    try:
        with engine.connect() as conn:
            result  = conn.execute(text(raw_sql))
            columns = list(result.keys())
            all_rows = [
                [str(v) if v is not None else '' for v in row]
                for row in result.fetchall()
            ]
    except Exception as exc:
        raise HTTPException(status_code=400,
            detail=f'Query execution failed: {exc}\n\nGenerated SQL:\n{raw_sql}')

    return {
        'sql':     raw_sql,
        'columns': columns,
        'preview': all_rows[:20],
        'total':   len(all_rows),
    }


# ── /download ─────────────────────────────────────────────────────────────────
@router.post('/download')
def download_export(req: DownloadRequest,
                    db: Session = Depends(get_db),
                    _:  models.User = Depends(require_admin)):
    """
    Re-validate the SQL from a previous /preview call and stream as a file.
    """
    import io, csv as _csv
    from fastapi.responses import StreamingResponse, Response

    sql = req.sql.strip().rstrip(';')

    if not _validate_select(sql):
        raise HTTPException(status_code=400, detail='Invalid or unsafe SQL.')

    try:
        with engine.connect() as conn:
            result  = conn.execute(text(sql))
            columns = list(result.keys())
            rows    = [
                [str(v) if v is not None else '' for v in row]
                for row in result.fetchall()
            ]
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f'Query execution failed: {exc}')

    fmt = (req.format or 'csv').lower()

    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500,
                detail='openpyxl not installed — run: pip install openpyxl')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AI Export'
        hfill = PatternFill('solid', fgColor='1E3A5F')
        hfont = Font(color='FFFFFF', bold=True)
        ws.append(columns)
        for cell in ws[1]:
            cell.font  = hfont
            cell.fill  = hfill
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            width = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 45)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="ai_export.xlsx"'})

    delimiter = '\t' if fmt == 'tsv' else ','
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=delimiter)
    writer.writerow(columns)
    writer.writerows(rows)
    ext  = 'tsv' if fmt == 'tsv' else 'csv'
    mime = 'text/tab-separated-values' if fmt == 'tsv' else 'text/csv'
    return Response(content=buf.getvalue(), media_type=mime,
        headers={'Content-Disposition': f'attachment; filename="ai_export.{ext}"'})
